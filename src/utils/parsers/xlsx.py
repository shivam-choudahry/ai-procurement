"""
Excel (XLSX / XLS) parser — converts spreadsheet content to readable plain text.
Each sheet is rendered as a Markdown-style table so the LLM can parse it easily.
"""

import io
import logging

logger = logging.getLogger(__name__)


def parse_xlsx(content: bytes) -> str:
    """
    Parse an Excel file (.xlsx / .xls) to plain text.

    Each worksheet is converted to a Markdown table.  Empty sheets are skipped.

    Args:
        content: Raw bytes of the Excel file

    Returns:
        Plain-text representation of all sheets
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl is not installed — cannot parse XLSX files. Run: pip install openpyxl")
        return "[XLSX parsing failed: openpyxl not installed]"

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"xlsx parser: failed to open workbook — {e}")
        return f"[XLSX parsing failed: {e}]"

    parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Drop fully-empty rows
        non_empty = [r for r in rows if any(cell is not None and str(cell).strip() for cell in r)]
        if not non_empty:
            continue

        parts.append(f"## Sheet: {sheet_name}\n")

        # Use first non-empty row as header
        header = [str(c) if c is not None else "" for c in non_empty[0]]
        parts.append("| " + " | ".join(header) + " |")
        parts.append("| " + " | ".join("---" for _ in header) + " |")

        for row in non_empty[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            # Pad row to header width
            while len(cells) < len(header):
                cells.append("")
            parts.append("| " + " | ".join(cells[: len(header)]) + " |")

        parts.append("")  # blank line between sheets

    wb.close()

    result = "\n".join(parts).strip()
    logger.info(f"xlsx parser: extracted {len(result)} chars from {len(wb.sheetnames)} sheet(s)")
    return result if result else "[XLSX file contained no readable data]"

